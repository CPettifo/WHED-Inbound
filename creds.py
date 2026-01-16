# This file will process the standardised credential data and convert it into the appropriate WHED Codes

import pymysql, cryptography, os, tempfile
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from collections import Counter

def main(masterlist_path, output_path):
    # Create a new dict of institutions
    insts = get_insts(output_path)

    # Get list of FOS Codes and FOS Levels / Display Categories from WHED (or spreadsheet)
    # Open connection to the WHED
    conn = whed_connect()
    cursor = conn.cursor()
    # test connection
    cursor.execute("SELECT GlobalID, OrgName FROM whed_org LIMIT 20;")

    # Open masterlist
    if not os.path.exists(masterlist_path):
        print(f"Masterlist not found at {masterlist_path}")
        return
    
    #Read list
    print("Opening masterlist be patient...", flush = True)
    wb = load_workbook(masterlist_path)

    # Open whed_levels sheet
    whed_levels = wb['whed_levels']
  

    # get list of whed credentials from whed_levels
    whed_creds = get_whed_creds(whed_levels, cursor)


    # get list of fos from the WHED
    whed_fos = get_whed_fos(cursor)

    # open courses sheet
    ext_cred = wb['ext_cred']

    processed = 0
    dual_qual_count = 0
    matched_creds = 0
    unmatched_creds = 0
    unmatched_creds_list = set()
    matched_fos = 0
    matched_fos_list = Counter()
    unmatched_fos = 0
    unmatched_fos_list = Counter()

    for inst in insts:
        # skip institition if it hasn't been categorised as "confirmed"
        if inst["status"] != "confirmed":
            continue
        # For each credential in sheet
        for row in ext_cred.iter_rows(min_row=2, values_only = True):

            # If credential's institution matches the one in the loop & is not expired, and is one of the credentials we keep in the WHED
            expired = str(row[2])
            if expired == "No" and inst["ext_id"] == str(row[0]):
                processed += 1

                course_level= str(row[3])
                course_name = str(row[4])
                fos_level_1 = str(row[5])
                fos_level_2 = str(row[6])
                fos_level_3 = str(row[7])
                
                # Check that the credential is one that we want to store in the WHED

                whed_level_codes = ["6B", "6C", "7A", "7B", "7C"]
                

                match = any(
                    whed_cred["cred_name"] == course_level and whed_cred["cred_level_code"] in whed_level_codes
                    for whed_cred in whed_creds
                )

                if match:
                    matched_creds += 1
                else:
                    unmatched_creds += 1
                    unmatched_creds_list.add(course_level)
                

                # The degree using information from the external data source
                ext_degree = {
                    "course_level":course_level,
                    "course_name": course_name,
                    "fos_level_1": fos_level_1,
                    "fos_level_2": fos_level_2,
                    "fos_level_3": fos_level_3,
                    "dual_qualification": str(row[22])
                }
                # increment variable if there is a dual qualification
                if ext_degree["dual_qualification"] == "Yes":
                    dual_qual_count += 1
                    continue
                
                # Match whed credential code
                cred_code = get_cred_code(whed_creds, ext_degree)

                # Attempt to match by fos_code
                fos_code, course_field = get_fos_code(whed_fos, ext_degree)
                if fos_code == "????":
                    unmatched_fos += 1
                    unmatched_fos_list[course_field] += 1
                else:
                    matched_fos += 1
                    matched_fos_list[course_field] += 1


                # Assign the degree (cred (bachelor, masters) + field of study (compsci, history)) to a variable
                degree = {
                    "org_id": inst["whed_id"],
                    "cred_code": cred_code,
                    "fos_code": fos_code,
                    "course_name": ext_degree["course_name"]
                }
                if (processed % 100 == 0):
                    print(f"Processed credentials: {processed}")


    print(f"List of matched fields of study ({matched_fos}/{processed}): \n")
    for fos, count in matched_fos_list.most_common():
        print(f"{fos}: {count}")

    # Columns to include in full output
    # Ext. Inst ID
    # WHED Inst ID (For now it is ok if this is empty)
    # CourseLevel
    # Ext. CredID
    # WHED FOSCode (if match for now ignore this)
    # Ext. FOS Levels
    # Full Course Name (degree["course_name")


    print(f"\nList of unmatched fields of study ({unmatched_fos}/{processed}): \n")
    for fos, count in unmatched_fos_list.most_common():
        print(f"{fos}: {count}")
    print(f"\nCredentials not in WHED List: {unmatched_creds}\nList: {unmatched_creds_list}\n-----------------------------------------------------------\n")
    print(f"\n\nTotal Records Processed: {processed}\n-----------------------------------------------------------\n\nMatched Credentials: {matched_creds}\nUnmatched Credentials: {unmatched_creds}")
    print(f"\n-----------------------------------------------------------\n\nMatched FOS: {matched_fos}\nUnmatched FOS: {unmatched_fos}\n\n-----------------------------------------------------------\n")
    print(f"There were {dual_qual_count} valid dual qualifications processed")
    exit

# Will test the connection to the db
def whed_test_connect(query):
    load_dotenv()
    timeout = 10
    connection = pymysql.connect(
        charset="utf8mb4",
        connect_timeout=timeout,
        cursorclass=pymysql.cursors.DictCursor,
        db=os.getenv("DB_NAME"),
        host=os.getenv("DB_HOST"),
        password=os.getenv("DB_PASSWORD"),
        read_timeout=timeout,
        port=int(os.getenv("DB_PORT", 3306)),
        user=os.getenv("DB_USER"),
        write_timeout=timeout,
    )
    try:
        cursor = connection.cursor()
        cursor.execute(query)
        print(cursor.fetchall())
    finally:
        connection.close()

# will return the conn for the database connection
def whed_connect():
    load_dotenv()
    timeout = 10
    connection = pymysql.connect(
        charset="utf8mb4",
        connect_timeout=timeout,
        cursorclass=pymysql.cursors.DictCursor,
        db=os.getenv("DB_NAME"),
        host=os.getenv("DB_HOST"),
        password=os.getenv("DB_PASSWORD"),
        read_timeout=timeout,
        port=int(os.getenv("DB_PORT", 3306)),
        user=os.getenv("DB_USER"),
        write_timeout=timeout,
    )
    return connection

# takes the current row of the credentials object and the whed_levels sheet to try match and return the credential code
def get_cred_code(whed_creds, ext_degree):
    
   


    return "1138"

def get_fos_code(whed_fos, ext_degree):
    # First we need to turn the string into substrings, i.e. Bachelor of Computer Science -> Computer Science
    course_name = ext_degree["course_name"]
    strings = []
    if " of " in course_name:
        strings = course_name.split(" of ")
    elif " in " in course_name:
        strings = course_name.split(" in ")
    else:
        strings = ["Placeholder", course_name]
    course_field = strings[1]

    match = any(
    whed_field["FOSDisplay"] == course_field
    for whed_field in whed_fos
    )

    if match:
        # get matching fos code, for now return 1111
        return 1111, course_field


    if not match:
        # Debug print
        print(f"--------------------------\nNot a match\nExternal Degree: {course_name}\nField Name: {course_field}\n--------------------------")
        return "????", course_field
    


    #TODO Match Field to appropriate whed FOS using the following hierarchy
    # If any of the FOS fields match, use that
        # Get WHED FOS Code and add it to a dict
    # If a shaved version of the credential name matches a WHED FOS field
        # Get WHED FOS Code and add it to a dict
    # If there is a fuzzy match
        # Add the cred to the "to be sorted" category, and add to a bucket
        # By bucket I mean basically to have all unsorted categories matched together, so there could potentially be 100 instances of a
        # non-matched field (e.g. Mobile Programming) that could then be categorised by a Data Officer at the end of the program
    #print("FOS Code Row: ", row)
    return "1111"

# Will return a dict of all institutions in the source list
def get_insts(output_path):
    # initialise insts list
    insts = []
    # open output file
    if not os.path.exists(output_path):
        print(f"Output not found, did you run the categorisation on the masterlist? path: {output_path}")
        exit()

    # Read output
    print("Opening output", flush = True)
    wb = load_workbook(output_path)
    
    # Open output sheet
    ws = wb['Sheet']

    # Loop through rows
    for row in ws.iter_rows(min_row=2, values_only = True):
        # create new inst using row info
        inst = {
            "whed_id": row[3],
            "whed_name": str(row[8]),
            "whed_name_eng": str(row[6]),
            "ext_id": str(row[1]),
            "ext_name": str(row[2]),
            "ext_trading": str(row[0]),
            "status": str(row[5]),
            "match_type": str(row[7])
            }
        insts.append(inst)

    return insts

# Returns a list of credentials in the WHED for country specified in source file.
def get_whed_creds(whed_levels, cursor):
    whed_creds = []
    for row in whed_levels.iter_rows(min_row=2, values_only = True):
        cred_name = str(row[0])
        cred_level_code = str(row[1])
        country_code = str(row[2])

        # get country ID
        cursor.execute(f"SELECT StateID FROM whed_state WHERE CountryCode = %s", (country_code,))
        result = cursor.fetchone()
        country_id = result["StateID"]

        # get CredID of credential using credential name and country ID
        cursor.execute(f"SELECT CredID FROM whed_cred WHERE StateID = %s", (country_id))
        result = cursor.fetchone()
        cred_id = result["CredID"]
                       

        whed_cred = {
            "cred_name": cred_name,
            "cred_level_code": cred_level_code,
            "cred_id": cred_id,
            "country_id": country_id
        }
        #print(whed_cred)
        whed_creds.append(whed_cred)

    return whed_creds

# Get a list of all fields of study in the WHED
def get_whed_fos(cursor):
    cursor.execute(f"SELECT FOSCode, FOSLevel1, FOSLevel2, FOSLevel3, FOSDisplay FROM whed_lex_fos ORDER BY FOSLevel1, FOSLevel2, FOSLevel3")
    results = cursor.fetchall()
    whed_fos = []

    for result in results:
        fos = {
            "FOSCode": result["FOSCode"],
            "FOSDisplay": result["FOSDisplay"],
            "FOSLevel1": result["FOSLevel1"],
            "FOSLevel2": result["FOSLevel2"],
            "FOSLevel3": result["FOSLevel3"]
        }
        whed_fos.append(fos)

    return whed_fos